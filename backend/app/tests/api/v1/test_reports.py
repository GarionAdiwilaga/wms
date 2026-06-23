import pytest
from datetime import datetime, timezone, timedelta
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session
from openpyxl import load_workbook
import io
import csv

from app.models.branch import Branch
from app.models.user import User
from app.models.item import Item
from app.models.category import Category
from app.models.supplier import Supplier
from app.models.uom import UOM
from app.models.inventory import InventoryTransaction, BranchStock
from app.models.transfer import Transfer, TransferLine
from app.models.audit_log import AuditLog
from app.core.security import create_access_token, get_password_hash

def get_auth_headers(user_id: int, role: str) -> dict[str, str]:
    token = create_access_token(user_id=user_id, role=role, token_version=1)
    return {"Authorization": f"Bearer {token}"}

@pytest.fixture(scope="function")
def test_data(db_session: Session):
    # Create Branches
    branch_a = Branch(code="BRA", name="Branch A", location="Location A", is_active=True)
    branch_b = Branch(code="BRB", name="Branch B", location="Location B", is_active=True)
    db_session.add_all([branch_a, branch_b])
    db_session.flush()

    # Create Users
    admin = User(
        username="super_admin_test",
        password_hash=get_password_hash("admin123"),
        full_name="Super Admin User",
        role="super_admin",
        branch_id=None,
        is_active=True
    )
    head_a = User(
        username="branch_head_a",
        password_hash=get_password_hash("head123"),
        full_name="Branch Head A",
        role="branch_head",
        branch_id=branch_a.branch_id,
        is_active=True
    )
    staff_a = User(
        username="staff_a",
        password_hash=get_password_hash("staff123"),
        full_name="Staff A",
        role="warehouse_staff",
        branch_id=branch_a.branch_id,
        is_active=True
    )
    db_session.add_all([admin, head_a, staff_a])
    db_session.flush()

    # Create Item Metadata
    cat = Category(code="CAT", name="Category A")
    sup = Supplier(code="SUP", name="Supplier A")
    uom = UOM(code="UOM", name="PCS")
    db_session.add_all([cat, sup, uom])
    db_session.flush()

    # Create Items
    item1 = Item(
        item_code="CAT-SUP-001",
        manual_code="001",
        name="Item 1",
        category_id=cat.category_id,
        supplier_id=sup.supplier_id,
        uom_id=uom.uom_id,
        minimum_stock=10,
        is_active=True
    )
    item2 = Item(
        item_code="CAT-SUP-002",
        manual_code="002",
        name="Item 2",
        category_id=cat.category_id,
        supplier_id=sup.supplier_id,
        uom_id=uom.uom_id,
        minimum_stock=50,
        is_active=True
    )
    db_session.add_all([item1, item2])
    db_session.flush()

    # Branch Stocks
    stock_a1 = BranchStock(branch_id=branch_a.branch_id, item_id=item1.item_id, quantity=5) # Low stock: 5 <= 10, shortage = 5
    stock_a2 = BranchStock(branch_id=branch_a.branch_id, item_id=item2.item_id, quantity=60) # Normal stock
    stock_b1 = BranchStock(branch_id=branch_b.branch_id, item_id=item1.item_id, quantity=15) # Normal stock
    db_session.add_all([stock_a1, stock_a2, stock_b1])
    db_session.flush()

    return {
        "branch_a": branch_a,
        "branch_b": branch_b,
        "admin": admin,
        "head_a": head_a,
        "staff_a": staff_a,
        "item1": item1,
        "item2": item2,
        "cat": cat,
        "sup": sup
    }

def test_stock_report_permissions(db_client: TestClient, test_data):
    # Warehouse staff is blocked
    headers = get_auth_headers(test_data["staff_a"].user_id, "warehouse_staff")
    response = db_client.get("/api/v1/reports/stock", headers=headers)
    assert response.status_code == 403

    # Branch Head is restricted to their branch
    headers = get_auth_headers(test_data["head_a"].user_id, "branch_head")
    response = db_client.get(f"/api/v1/reports/stock?branch_id={test_data['branch_b'].branch_id}", headers=headers)
    assert response.status_code == 403

def test_stock_report_super_admin(db_client: TestClient, test_data):
    headers = get_auth_headers(test_data["admin"].user_id, "super_admin")
    response = db_client.get("/api/v1/reports/stock?search=CAT-SUP", headers=headers)
    assert response.status_code == 200
    content = response.json()
    assert content["total"] == 3
    
    # Verify filtering by category
    response = db_client.get(f"/api/v1/reports/stock?category_id={test_data['cat'].category_id}&search=CAT-SUP", headers=headers)
    assert response.status_code == 200
    assert response.json()["total"] == 3

    # Verify search functionality
    response = db_client.get("/api/v1/reports/stock?search=002", headers=headers)
    assert response.status_code == 200
    assert response.json()["total"] == 1
    assert response.json()["items"][0]["item_code"] == "CAT-SUP-002"

def test_low_stock_report(db_client: TestClient, test_data):
    headers = get_auth_headers(test_data["admin"].user_id, "super_admin")
    response = db_client.get("/api/v1/reports/low-stock?search=CAT-SUP", headers=headers)
    assert response.status_code == 200
    content = response.json()
    assert content["total"] == 1 # Only Item 1 at Branch A is low stock
    item = content["items"][0]
    assert item["item_code"] == "CAT-SUP-001"
    assert item["branch_name"] == "Branch A"
    assert item["quantity"] == 5
    assert item["minimum_stock"] == 10
    assert item["shortage"] == 5

def test_item_history_report(db_client: TestClient, db_session: Session, test_data):
    # Populate historical transactions
    t1 = datetime.now(timezone.utc) - timedelta(days=3)
    t2 = datetime.now(timezone.utc) - timedelta(days=2)
    tx1 = InventoryTransaction(
        item_id=test_data["item1"].item_id,
        branch_id=test_data["branch_a"].branch_id,
        transaction_type="IN",
        quantity=10,
        reference_type="initial_load",
        created_by=test_data["admin"].user_id,
        created_at=t1
    )
    tx2 = InventoryTransaction(
        item_id=test_data["item1"].item_id,
        branch_id=test_data["branch_a"].branch_id,
        transaction_type="OUT",
        quantity=3,
        reference_type="outbound",
        created_by=test_data["admin"].user_id,
        created_at=t2
    )
    db_session.add_all([tx1, tx2])
    db_session.flush()

    headers = get_auth_headers(test_data["admin"].user_id, "super_admin")
    response = db_client.get(f"/api/v1/reports/item-history/{test_data['item1'].item_id}", headers=headers)
    assert response.status_code == 200
    content = response.json()
    assert content["total"] == 2
    assert content["items"][0]["transaction_type"] == "OUT"
    assert content["items"][1]["transaction_type"] == "IN"

def test_inventory_movement_report_balance_after(db_client: TestClient, db_session: Session, test_data):
    # Populating sequential movements to test running balance after date filters
    t1 = datetime(2026, 1, 1, 10, 0, tzinfo=timezone.utc)
    t2 = datetime(2026, 1, 2, 10, 0, tzinfo=timezone.utc)
    t3 = datetime(2026, 1, 3, 10, 0, tzinfo=timezone.utc)

    tx1 = InventoryTransaction(
        item_id=test_data["item1"].item_id,
        branch_id=test_data["branch_a"].branch_id,
        transaction_type="IN",
        quantity=10,
        reference_type="initial_load",
        created_by=test_data["admin"].user_id,
        created_at=t1
    )
    tx2 = InventoryTransaction(
        item_id=test_data["item1"].item_id,
        branch_id=test_data["branch_a"].branch_id,
        transaction_type="OUT",
        quantity=3,
        reference_type="outbound",
        created_by=test_data["admin"].user_id,
        created_at=t2
    )
    tx3 = InventoryTransaction(
        item_id=test_data["item1"].item_id,
        branch_id=test_data["branch_a"].branch_id,
        transaction_type="IN",
        quantity=5,
        reference_type="stock_in",
        created_by=test_data["admin"].user_id,
        created_at=t3
    )
    db_session.add_all([tx1, tx2, tx3])
    db_session.flush()

    headers = get_auth_headers(test_data["admin"].user_id, "super_admin")
    
    # 1. Verify movements list and balance_after ordering
    response = db_client.get(f"/api/v1/reports/movements?branch_id={test_data['branch_a'].branch_id}", headers=headers)
    assert response.status_code == 200
    content = response.json()
    assert content["total"] == 3
    # Check that latest transaction is first, but running balance calculated chronologically
    items = content["items"]
    assert items[0]["transaction_type"] == "IN"
    assert items[0]["balance_after"] == 12 # 10 - 3 + 5
    assert items[1]["transaction_type"] == "OUT"
    assert items[1]["balance_after"] == 7 # 10 - 3
    assert items[2]["transaction_type"] == "IN"
    assert items[2]["balance_after"] == 10 # 10

    # 2. Filter with start_date = 2026-01-02. The first transaction is excluded,
    # but the running balance for tx2 and tx3 must still be correct (7 and 12, not -3 and 2)
    response = db_client.get(
        f"/api/v1/reports/movements?branch_id={test_data['branch_a'].branch_id}&start_date=2026-01-02T00:00:00Z",
        headers=headers
    )
    assert response.status_code == 200
    content = response.json()
    assert content["total"] == 2
    assert content["items"][0]["balance_after"] == 12
    assert content["items"][1]["balance_after"] == 7

def test_transfer_variance_report(db_client: TestClient, db_session: Session, test_data):
    # Create Received Transfer with Variance
    tr = Transfer(
        transfer_number="TRF-20260101-001",
        source_branch_id=test_data["branch_a"].branch_id,
        dest_branch_id=test_data["branch_b"].branch_id,
        status="received",
        created_by=test_data["admin"].user_id,
        received_by=test_data["admin"].user_id,
        received_at=datetime.now(timezone.utc)
    )
    db_session.add(tr)
    db_session.flush()

    line = TransferLine(
        transfer_id=tr.transfer_id,
        item_id=test_data["item1"].item_id,
        sent_quantity=10,
        received_quantity=8,
        variance_reason="Pecah",
        variance_notes="Rusak saat transit"
    )
    db_session.add(line)
    db_session.flush()

    headers = get_auth_headers(test_data["admin"].user_id, "super_admin")
    response = db_client.get("/api/v1/reports/transfer-variance", headers=headers)
    assert response.status_code == 200
    content = response.json()
    assert content["total"] == 1
    assert content["items"][0]["variance"] == 2
    assert content["items"][0]["variance_reason"] == "Pecah"
    
    # Check Summary Metrics
    summary = content["summary"]
    assert summary["total_transfers"] == 1
    assert summary["transfers_with_variance"] == 1
    assert summary["total_lost_units"] == 2

def test_audit_log_report(db_client: TestClient, db_session: Session, test_data):
    # Add Audit Logs
    log1 = AuditLog(
        user_id=test_data["head_a"].user_id,
        action="CREATE",
        entity_type="items",
        entity_id=1,
        old_values=None,
        new_values={"name": "New Item"},
        ip_address="127.0.0.1"
    )
    db_session.add(log1)
    db_session.flush()

    # Head of Branch A queries logs
    headers = get_auth_headers(test_data["head_a"].user_id, "branch_head")
    response = db_client.get("/api/v1/reports/audit-logs", headers=headers)
    assert response.status_code == 200
    content = response.json()
    assert content["total"] == 1
    assert content["items"][0]["operator_name"] == "Branch Head A"

def test_csv_export_streaming(db_client: TestClient, test_data):
    headers = get_auth_headers(test_data["admin"].user_id, "super_admin")
    response = db_client.get("/api/v1/reports/stock?export=csv&search=CAT-SUP", headers=headers)
    assert response.status_code == 200
    assert response.headers["content-type"] == "text/csv; charset=utf-8"
    
    content = response.text
    csv_reader = csv.reader(io.StringIO(content))
    csv_rows = list(csv_reader)
    
    assert len(csv_rows) == 4 # Header + 3 rows
    assert csv_rows[0] == ["Cabang", "Kode Barang", "Nama Barang", "Kategori", "Supplier", "Stok", "Min Stok"]

def test_xlsx_export_streaming(db_client: TestClient, test_data):
    headers = get_auth_headers(test_data["admin"].user_id, "super_admin")
    response = db_client.get("/api/v1/reports/stock?export=xlsx&search=CAT-SUP", headers=headers)
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Read the streamed bytes back as an openpyxl workbook
    wb = load_workbook(io.BytesIO(response.content))
    sheet = wb.active
    assert sheet.max_row == 4 # Header + 3 rows
    assert sheet.cell(row=1, column=1).value == "Cabang"
